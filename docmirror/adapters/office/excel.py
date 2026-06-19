# Copyright (c) 2026 ValueMap Global and contributors. All rights reserved.
# Author: Adam Lin <adamlin@valuemapglobal.com>
#
# This source code is licensed under the Apache 2.0 license found in the
# LICENSE file in the root directory of this source tree.

"""
Excel Adapter — .xlsx / .csv → ParseResult (Cell-level precision)
=================================================================

Native ``.xlsx`` via openpyxl; ``.csv`` delegated to StructuredAdapter.
Legacy ``.xls`` is transcoded to ``.xlsx`` upstream by ``TranscodingGate`` (FCR).
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime, time
from pathlib import Path

from docmirror.framework.base import BaseParser
from docmirror.models.entities.domain import BaseResult

logger = logging.getLogger(__name__)

# Pattern for currency-like values: optional sign, digits with commas, decimal part
_CURRENCY_RE = re.compile(r"^[¥$€£₹]?\s*-?\d{1,3}(,\d{3})*(\.\d+)?$")


def _classify_cell(value) -> CellValue:
    """Convert an openpyxl cell value to a typed CellValue."""
    from docmirror.models.entities.parse_result import CellValue, DataType

    if value is None:
        return CellValue(text="", data_type=DataType.EMPTY)

    if isinstance(value, (int, float)):
        text = str(value)
        # Detect currency-like (large numbers with 2 decimal places)
        if isinstance(value, float) and round(value, 2) == value:
            return CellValue(
                text=text,
                cleaned=text,
                numeric=float(value),
                data_type=DataType.CURRENCY,
            )
        return CellValue(
            text=text,
            cleaned=text,
            numeric=float(value),
            data_type=DataType.NUMBER,
        )

    if isinstance(value, (datetime, date)):
        text = str(value)
        return CellValue(text=text, data_type=DataType.DATE)

    if isinstance(value, time):
        text = str(value)
        return CellValue(text=text, data_type=DataType.TEXT)

    # String values — try to detect numeric/currency strings
    text = str(value).strip()
    if not text:
        return CellValue(text="", data_type=DataType.EMPTY)

    # Try to parse currency-like strings: "15,000.00", "$1,234.56", "¥100"
    if _CURRENCY_RE.match(text):
        cleaned = re.sub(r"[¥$€£₹,\s]", "", text)
        try:
            numeric = float(cleaned)
            return CellValue(
                text=text,
                cleaned=cleaned,
                numeric=numeric,
                data_type=DataType.CURRENCY,
            )
        except ValueError:
            pass

    return CellValue(text=text, data_type=DataType.TEXT)


class ExcelAdapter(BaseParser):
    """Excel (.xlsx / .csv) adapter — openpyxl or CSV delegate; .xls via FCR transcode."""

    async def to_parse_result(self, file_path: Path, **kwargs) -> ParseResult:
        """
        Parse .xlsx directly into ParseResult with Cell-level precision.

        Each cell preserves its original data type (NUMBER, CURRENCY, DATE, TEXT)
        and numeric value, enabling downstream consumers to work with typed data
        instead of re-parsing strings.
        """
        import openpyxl

        from docmirror.models.entities.parse_result import (
            PageContent,
            ParseResult,
            ParserInfo,
            TableBlock,
            TableRow,
            TextBlock,
            TextLevel,
        )

        from docmirror.models.errors import build_failure_result

        ext = file_path.suffix.lower()
        if ext not in (".xlsx", ".csv"):
            return build_failure_result(
                "FORMAT_REQUIRES_CONVERTER",
                f"ExcelAdapter requires .xlsx or .csv; got {ext}. "
                "Legacy .xls is transcoded by the extraction pipeline.",
                file_path=str(file_path),
                file_type="excel",
            )

        if ext == ".csv":
            # Tabular export (bank statement, regulatory filing) — same table
            # semantics as Excel; CSV parsing delegated to StructuredAdapter.
            from docmirror.adapters.data.structured import StructuredAdapter

            logger.info(f"[ExcelAdapter] Delegating CSV to StructuredAdapter: {file_path}")
            return await StructuredAdapter().to_parse_result(file_path, **kwargs)

        logger.info(f"[ExcelAdapter] Cell-level extraction for: {file_path}")
        wb = openpyxl.load_workbook(str(file_path), data_only=True)

        pages: list[PageContent] = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]

            # Collect all rows with typed CellValues
            all_rows: list[list] = []
            for row in sheet.iter_rows(values_only=True):
                typed_cells = [_classify_cell(c) for c in row]
                # Skip completely empty rows
                from docmirror.models.entities.parse_result import DataType

                if all(c.data_type == DataType.EMPTY for c in typed_cells):
                    continue
                all_rows.append(typed_cells)

            if not all_rows:
                continue

            # First row as headers, rest as data rows
            header_cells = all_rows[0]
            headers = [c.text for c in header_cells]

            data_rows = [
                TableRow(
                    cells=cells,
                    row_type="data",
                    source_page=idx,
                )
                for cells in all_rows[1:]
            ]

            table = TableBlock(
                table_id=f"sheet{idx}_{sheet_name}",
                headers=headers,
                rows=data_rows,
                page=idx,
                caption=sheet_name,
            )

            page = PageContent(
                page_number=idx,
                texts=[TextBlock(content=f"Sheet: {sheet_name}", level=TextLevel.H2)],
                tables=[table],
            )
            pages.append(page)

        return ParseResult(
            pages=pages,
            parser_info=ParserInfo(
                parser_name="ExcelAdapter",
                page_count=len(pages),
                overall_confidence=1.0,
            ),
        )

    async def to_base_result(self, file_path: Path) -> BaseResult:
        """
        Fallback: Parse .xlsx into BaseResult via ParseResultBridge.

        Delegates to ``to_parse_result()`` then converts down to BaseResult,
        preserving the Cell-level precision extraction path.
        """
        from docmirror.models.construction.parse_result_bridge import ParseResultBridge

        pr = await self.to_parse_result(file_path)
        return ParseResultBridge.to_base_result(pr)
